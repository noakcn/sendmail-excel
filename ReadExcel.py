# -*- coding: UTF-8 -*-
from openpyxl import Workbook
from openpyxl import load_workbook

wb = load_workbook('example.xlsx')
print(wb.sheetnames)
for row in wb:
    for cell in row:
        for v in cell:
            print(v.value)
            # 读取email地址
            if v.value == 'email':
                print(v.value)

print("读取完毕...")

